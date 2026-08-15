#!/usr/bin/env python3
"""Regenerate the standalone JSON and Excel plant inventories from data.js."""

from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    bundled_python = (
        Path.home()
        / ".cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
    )
    if bundled_python.exists() and Path(sys.executable) != bundled_python:
        os.execv(str(bundled_python), [str(bundled_python), *sys.argv])
    raise SystemExit(
        "openpyxl is required to generate the Excel guide. "
        "Run this through the Codex document runtime or install openpyxl."
    ) from exc


ROOT = Path(__file__).resolve().parent
JSON_PATH = ROOT / "data" / "plants.json"
XLSX_PATH = ROOT / "Oak_Lodge_Garden_Plant_Guide.xlsx"

JXA_EXPORT = r'''
ObjC.import("Foundation");
function readText(path) {
  const value = $.NSString.stringWithContentsOfFileEncodingError(
    $(path).stringByStandardizingPath,
    $.NSUTF8StringEncoding,
    null
  );
  if (!value) throw new Error(`Cannot read ${path}`);
  return ObjC.unwrap(value);
}
function run(argv) {
  const window = {};
  eval(readText(`${argv[0]}/data.js`));
  return JSON.stringify(window.OAK.PLANTS);
}
'''


def load_plants() -> dict[str, list[dict]]:
    result = subprocess.run(
        ["/usr/bin/osascript", "-l", "JavaScript", "-e", JXA_EXPORT, str(ROOT)],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


def write_json(plants: dict[str, list[dict]]) -> None:
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    JSON_PATH.write_text(
        json.dumps(plants, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_workbook(plants: dict[str, list[dict]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plant Guide"
    headers = [
        "Location",
        "English Name",
        "Latin Name",
        "Light Requirements",
        "Watering & Soil",
        "General Care",
        "Seasonal Changes (What happens over the year)",
    ]
    sheet.append(headers)

    for location, records in plants.items():
        for plant in records:
            sheet.append([
                location,
                plant.get("name", ""),
                plant.get("latin", ""),
                plant.get("light", ""),
                plant.get("water", ""),
                plant.get("care", ""),
                plant.get("seasonal", ""),
            ])

    header_fill = PatternFill("solid", fgColor="355E3B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [24, 34, 38, 36, 42, 64, 64]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.print_title_rows = "1:1"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    workbook.save(XLSX_PATH)


def verify(plants: dict[str, list[dict]]) -> int:
    expected = [
        (location, plant.get("name", ""), plant.get("latin", ""))
        for location, records in plants.items()
        for plant in records
    ]
    exported = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    actual_json = [
        (location, plant.get("name", ""), plant.get("latin", ""))
        for location, records in exported.items()
        for plant in records
    ]
    if actual_json != expected:
        raise SystemExit("JSON verification failed: exported plant rows differ from data.js")

    workbook = load_workbook(XLSX_PATH, read_only=True, data_only=False)
    sheet = workbook["Plant Guide"]
    actual_xlsx = [tuple(row) for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True)]
    if actual_xlsx != expected:
        raise SystemExit("Excel verification failed: location/name/Latin rows differ from data.js")
    return len(expected)


def main() -> None:
    plants = load_plants()
    write_json(plants)
    write_workbook(plants)
    count = verify(plants)
    print(f"Plant exports ready: {count} records across {len(plants)} groups.")


if __name__ == "__main__":
    main()
